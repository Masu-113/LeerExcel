from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import openpyxl.styles as styles

def crear_imagen_con_formato_excel(archivo_excel, hoja_nombre, archivo_imagen):
    # Cargar el libro de trabajo y la hoja
    workbook = load_workbook(filename=archivo_excel)
    sheet = workbook[hoja_nombre]

    # Obtener dimensiones de la hoja
    max_row = sheet.max_row
    max_column = sheet.max_column

    # Definir fuente y tamaño de fuente
    font_size = 12
    try:
        font = ImageFont.truetype("Arial.ttf", font_size) # Reemplaza con la ruta a tu fuente
    except IOError:
        font = ImageFont.load_default() # Usa la fuente por defecto si no se encuentra Arial

    # Calcular el ancho y alto de cada celda (ajustar si es necesario)
    cell_width = 120
    cell_height = 25

    # Calcular el ancho y alto total de la imagen
    image_width = cell_width * max_column
    image_height = cell_height * max_row

    # Crear una imagen con fondo blanco
    image = Image.new('RGB', (image_width, image_height), 'white')
    draw = ImageDraw.Draw(image)

    # Iterar sobre las celdas y dibujar su contenido y formato
    for row in range(1, max_row + 1):
        for col in range(1, max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell_value = cell.value

            # Obtener posición de la celda
            x = (col - 1) * cell_width
            y = (row - 1) * cell_height


            # Dibujar el fondo de la celda
            draw.rectangle((x, y, x + cell_width, y + cell_height), fill='white')

            # Dibujar el texto
            if cell_value is not None:
                draw.text((x + 5, y + 5), str(cell_value), font=font, fill='black')

            # Dibujar bordes (opcional, solo para visualización)
                draw.rectangle((x, y, x + cell_width, y + cell_height), outline="black")

    # Guardar la imagen
    image.save(archivo_imagen)
    print(f"Imagen guardada en: {archivo_imagen}")

# Ejemplo de uso:
archivo_excel = r'C:\Users\msuarez\Documents\TestExcel.xlsx'
hoja_nombre = 'Sheet1' 
archivo_imagen = r'C:\Users\msuarez\source\repos\LeerExcel\prueba\page_1.jpg'

crear_imagen_con_formato_excel(archivo_excel, hoja_nombre, archivo_imagen)